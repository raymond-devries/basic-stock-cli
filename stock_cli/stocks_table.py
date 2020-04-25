import yfinance as yf
from openpyxl.utils import get_column_letter
from ta import volume, momentum, volatility, trend
import pandas as pd
import datetime
from dateutil import parser
from dateutil.relativedelta import relativedelta

import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font, Alignment


class StocksTable:
    """
    :param
    stocks: list
        A list of tickers as strings
    period : str
        Valid periods: #d, #w, #m, #y, max
    interval : str
        Valid intervals: 1m,2m,5m,15m,30m,60m,90m,1h,1d,5d,1wk,1mo,3mo
        Intraday data cannot extend last 60 days
    end: str
        Download end date string (YYYY-MM-DD) or _datetime.
        Default is now
    """

    def __init__(self, stocks: list, period="1y", interval="1d", end="Today"):
        self._stocks = stocks
        self._period = period

        self._end = self._get_end_date(end)

        self._start = self._get_start_date(self._period, self._end)

        self._interval = interval

        self._requested_data = self._get_requested_data(
            self._stocks, self._interval, self._start, self._end
        )
        self._data_52_weeks = self._get_data_52_weeks(self._stocks, self._end)

        self._table = self._process_data(self._requested_data, self._data_52_weeks)

    def _get_requested_data(
            self, stocks: list, interval: str, start: datetime.date, end: datetime.date
    ) -> pd.DataFrame:
        end = self._end + relativedelta(days=1)
        print("Getting requested data...")
        stocks_joined = ",".join(stocks)
        data = yf.download(stocks_joined, interval=interval, start=start, end=end)
        if len(stocks) == 1:
            data.columns = [data.columns, len(data.columns) * [stocks[0].upper()]]

        return data

    def _get_data_52_weeks(self, stocks: list, end: datetime.date) -> pd.DataFrame:
        print("Getting 52 week low and high data...")
        end = self._end + relativedelta(days=1)
        stocks_joined = ",".join(stocks)
        start = self._get_start_date("1y", end)
        data = yf.download(stocks_joined, interval="3mo", start=start, end=end)
        if len(stocks) == 1:
            data.columns = [data.columns, len(data.columns) * [stocks[0].upper()]]

        return data

    @staticmethod
    def _get_end_date(date: str) -> datetime.date:
        if date == "Today":
            return datetime.date.today()
        else:
            return parser.parse(date).date()

    @staticmethod
    def _get_start_date(period: str, end_date: datetime.date) -> datetime.date:
        period = period.strip()

        if period == "max":
            return datetime.date(1900, 1, 1)

        try:
            period_units = period[-1]
            period_length = int(period[:-1])
        except ValueError:
            raise ValueError("Invalid period type")

        if period_units == "d":
            delta = relativedelta(days=period_length)
        elif period_units == "w":
            delta = relativedelta(weeks=period_length)
        elif period_units == "m":
            delta = relativedelta(months=period_length)
        elif period_units == "y":
            delta = relativedelta(years=period_length)
        else:
            raise ValueError("Invalid period type")

        return end_date - delta

    def _process_data(
            self, data: pd.DataFrame, data_52_weeks: pd.DataFrame
    ) -> pd.DataFrame:
        price = self._calculate_price(data)
        low_52_weeks = self._calculate_52_week_low(data_52_weeks)
        high_52_weeks = self._calculate_52_week_high(data_52_weeks)
        volume = self._calculate_volume(data)

        ema_11, ema_22 = self._calculate_ema(data, 11), self._calculate_ema(data, 22)
        ema_11.name, ema_22.name = "EMA 11 Day", "EMA 22 Day"

        rsi = self._calculate_rsi(data)

        force_index_2, force_index_13 = (
            self._calculate_force_index(data, 2),
            self._calculate_force_index(data, 13),
        )
        force_index_2.name, force_index_13.name = (
            "Force Index 2 Day",
            "Force Index 13 Day",
        )

        (
            macd,
            macd_signal,
            macd_histogram,
            macd_histogram_pct_change,
        ) = self._calculate_macd(data)

        (
            keltner_band_high_band_20_3,
            keltner_band_lower_band_20_3,
        ) = self._calculate_keltner_bands(data, 20, 3)
        keltner_band_high_band_20_3.name, keltner_band_lower_band_20_3.name = (
            "Keltner High Band",
            "Keltner Low Band",
        )

        return pd.concat(
            [
                price,
                low_52_weeks,
                high_52_weeks,
                volume,
                ema_11,
                ema_22,
                rsi,
                force_index_2,
                force_index_13,
                keltner_band_high_band_20_3,
                keltner_band_lower_band_20_3,
                macd,
                macd_signal,
                macd_histogram,
                macd_histogram_pct_change,
            ],
            axis=1,
        )

    @staticmethod
    def _calculate_price(data: pd.DataFrame) -> pd.Series:
        price = data["Adj Close"].iloc[-1]
        price.name = "Price"
        return price

    @staticmethod
    def _calculate_52_week_low(data: pd.DataFrame) -> pd.Series:
        low_52_weeks = data["Low"].min()
        low_52_weeks.name = "52 Week low"
        return low_52_weeks

    @staticmethod
    def _calculate_52_week_high(data: pd.DataFrame) -> pd.Series:
        high_52_weeks = data["High"].max()
        high_52_weeks.name = "52 Week high"
        return high_52_weeks

    @staticmethod
    def _calculate_volume(data: pd.DataFrame) -> pd.Series:
        volume = data["Volume"].iloc[-1]
        volume.name = "Volume"
        return volume

    @staticmethod
    def _calculate_ema(data: pd.DataFrame, span: int) -> pd.Series:
        ema = data["Adj Close"].ewm(span=span, adjust=False).mean().iloc[-1]
        return ema

    @staticmethod
    def _calculate_rsi(data: pd.DataFrame) -> pd.Series:
        rsi = data["Adj Close"].apply(momentum.rsi).iloc[-1]
        rsi.name = "RSI"
        return rsi

    @staticmethod
    def _calculate_force_index(data: pd.DataFrame, period: int) -> pd.Series:
        data = data.swaplevel(0, 1, 1)
        return (
            data.groupby(level=0, axis=1)
                .apply(
                lambda stock_data: volume.force_index(
                    stock_data.xs("Close", level=1, axis=1).squeeze(),
                    stock_data.xs("Volume", level=1, axis=1).squeeze(),
                    n=period,
                )
            )
                .iloc[-1]
        )

    @staticmethod
    def _calculate_macd(data: pd.DataFrame) -> tuple:
        macd_data = data["Adj Close"].apply(trend.MACD)

        macd_line = macd_data.apply(lambda x: x.macd()).iloc[:, -1]
        macd_signal = macd_data.apply(lambda x: x.macd_signal()).iloc[:, -1]

        macd_histogram_data = macd_data.apply(lambda x: x.macd_diff()).transpose()
        macd_histogram = macd_histogram_data.iloc[-1]
        macd_histogram_pct_change = macd_histogram_data.pct_change().iloc[-1]

        (
            macd_line.name,
            macd_signal.name,
            macd_histogram.name,
            macd_histogram_pct_change.name,
        ) = ("MACD", "MACD Signal", "MACD Histogram", "MACD Histogram Change")

        return macd_line, macd_histogram, macd_signal, macd_histogram_pct_change

    def _calculate_keltner_bands(
            self, data: pd.DataFrame, period: int, multiplier
    ) -> tuple:
        ema = self._calculate_ema(data, period)

        data = data.swaplevel(0, 1, 1)
        atr = (
            data.groupby(level=0, axis=1)
                .apply(
                lambda stock_data: volatility.average_true_range(
                    stock_data.xs("High", level=1, axis=1).squeeze(),
                    stock_data.xs("Low", level=1, axis=1).squeeze(),
                    stock_data.xs("Close", level=1, axis=1).squeeze(),
                    n=period,
                )
            )
                .iloc[-1]
        )

        upper_band = ema + multiplier * atr
        lower_band = ema - multiplier * atr

        return upper_band, lower_band

    def get_dataframe(self):
        return self._table

    def export_excel(self, name: str):
        workbook = openpyxl.Workbook()
        sheet = workbook.active

        sheet.merge_cells("A1:F1")
        sheet["A1"] = self._end
        sheet["A1"].font = Font(size=20)
        sheet["A1"].alignment = Alignment("center")
        sheet.row_dimensions[1].height = 25
        sheet.row_dimensions[2].height = 45
        sheet.row_dimensions[2].alignment = Alignment(wrap_text=True)
        sheet.freeze_panes = "B3"

        for row in dataframe_to_rows(self._table, index=True, header=True):
            sheet.append(row)

        offset = 2
        for i in range(offset, len(self._table.columns) + offset):
            column_letter = get_column_letter(i)
            sheet.column_dimensions[column_letter].width = 18
            sheet.column_dimensions[column_letter].number_format = "#,##0.00"

        workbook.save(name + " " + str(self._end) + ".xlsx")
